import pandas as pd
from openpyxl import load_workbook

def record_it(filename="", data=[], columns=[]):
    df1 = pd.DataFrame(data, columns=columns)
    try:
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        for sheetname in writer.sheets:
            if writer.sheets[sheetname].max_row > 1:
                wear_header = False
            else:
                wear_header = True
            df1.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,
                         header=wear_header)
        writer.save()
    except FileNotFoundError:
        df1.to_excel(filename, index=False)

    if data == [] or columns == []:
        return False
    return True
